import re
import math
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import List, Dict, Any
from ..config import COLUNAS_IMPORTANTES


def parse_price(value) -> float:
    """Converte um preço em diversos formatos para float; NaN se não for possível.

    Aceita, por exemplo: 150, 150.5, "150,50", "R$ 1.060,00" (padrão BR),
    "1,060.00" (padrão US), "N/A", "", None.
    """
    if value is None:
        return float('nan')
    if isinstance(value, (int, float)):
        v = float(value)
        return float('nan') if math.isnan(v) else v

    s = str(value).strip()
    # mantém apenas dígitos, ponto, vírgula e sinal
    s = re.sub(r'[^0-9.,\-]', '', s)
    if not s or s in ('-', '.', ','):
        return float('nan')

    if ',' in s and '.' in s:
        # Se a vírgula vem depois do ponto -> decimal é a vírgula (BR): 1.060,00
        # Caso contrário -> decimal é o ponto (US): 1,060.00
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        # só vírgula -> separador decimal BR: 222,00 -> 222.00
        s = s.replace(',', '.')
    # só ponto (ou nenhum) -> já está em formato float

    try:
        return float(s)
    except ValueError:
        return float('nan')

class DataManager:
    def __init__(self):
        self.df = None
        self.file_path = None
    
    def load_excel(self, file_path: str):
        """Carrega arquivo Excel"""
        try:
            self.df = pd.read_excel(file_path)
            self.file_path = file_path
            return True
        except Exception as e:
            raise Exception(f"Erro ao carregar Excel: {str(e)}")
    
    def save_excel(self, file_path: str = None):
        """Salva DataFrame no Excel com formatação"""
        if self.df is None:
            raise ValueError("Nenhum dado para salvar")
        
        save_path = file_path or self.file_path
        if not save_path:
            raise ValueError("Caminho do arquivo não especificado")
        
        self.df.to_excel(save_path, index=False, engine='openpyxl')
        
        wb = load_workbook(save_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, col_name in enumerate(self.df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col_name in COLUNAS_IMPORTANTES:
                cell.fill = yellow_fill
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(save_path)
        return save_path
    
    def add_rows(self, data: List[Dict[str, Any]]):
        """Adiciona novas linhas ao DataFrame"""
        if self.df is None:
            self.df = pd.DataFrame(data)
        else:
            new_df = pd.DataFrame(data)
            self.df = pd.concat([self.df, new_df], ignore_index=True)
    
    def get_dataframe(self):
        """Retorna o DataFrame atual"""
        return self.df
    
    def create_comparison(self, output_path: str):
        """Cria planilha de comparação de preços.

        Ordena por PRODUTO e depois por PREÇO (numérico, tolerante a formatos
        BR/US) e destaca em verde a linha de menor preço de cada produto.
        Tolera colunas ausentes: usa as colunas importantes que existirem,
        exigindo ao menos PRODUTO e PREÇO.
        """
        if self.df is None or self.df.empty:
            raise ValueError("Nenhum dado para comparar")

        # Exige as colunas mínimas para uma comparação de preços
        obrigatorias = ['PRODUTO', 'PREÇO']
        faltando = [c for c in obrigatorias if c not in self.df.columns]
        if faltando:
            raise ValueError(
                "Colunas obrigatórias ausentes para a comparação: "
                + ", ".join(faltando)
                + ".\nColunas encontradas: " + ", ".join(map(str, self.df.columns))
            )

        # Usa as colunas importantes que existirem, na ordem padrão
        cols = [c for c in COLUNAS_IMPORTANTES if c in self.df.columns]
        comparison_df = self.df[cols].copy()

        # Preço robusto (aceita "1.060,00", "R$ 222,00", "N/A", 150.5, ...)
        comparison_df['PREÇO'] = comparison_df['PREÇO'].map(parse_price)

        comparison_df = comparison_df.sort_values(
            by=['PRODUTO', 'PREÇO'], na_position='last'
        ).reset_index(drop=True)

        comparison_df.to_excel(output_path, index=False, engine='openpyxl')

        wb = load_workbook(output_path)
        ws = wb.active

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        n_cols = len(cols)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Coluna do PRODUTO (posição dinâmica) para detectar troca de produto
        produto_col = cols.index('PRODUTO') + 1
        current_product = None
        for row_idx in range(2, ws.max_row + 1):
            product_cell = ws.cell(row=row_idx, column=produto_col)
            if product_cell.value != current_product:
                current_product = product_cell.value
                for col_idx in range(1, n_cols + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        wb.save(output_path)
        return output_path
    
    def clear_data(self):
        """Limpa todos os dados"""
        self.df = None
        self.file_path = None
